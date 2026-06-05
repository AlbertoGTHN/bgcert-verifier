"""
Report Generation Service
Generates PDF, Excel, and CSV reports for certificate validation results.
"""
import os
import csv
import io
from datetime import datetime
from typing import List, Optional, Any
from pathlib import Path

from loguru import logger

from app.config import settings
from app.models.certificate import Certificate, ValidationStatus


STATUS_LABELS = {
    ValidationStatus.VERIFIED_AUTHENTIC: "VERIFIED AUTHENTIC",
    ValidationStatus.FAILED_FRAUDULENT: "FAILED / POSSIBLY FRAUDULENT",
    ValidationStatus.TECHNICAL_ISSUE: "TECHNICAL ISSUE",
    ValidationStatus.PENDING: "PENDING",
    ValidationStatus.PROCESSING: "PROCESSING",
    ValidationStatus.ERROR: "ERROR",
}

STATUS_COLORS_RGB = {
    ValidationStatus.VERIFIED_AUTHENTIC: (0, 150, 80),
    ValidationStatus.FAILED_FRAUDULENT: (200, 40, 40),
    ValidationStatus.TECHNICAL_ISSUE: (220, 160, 0),
    ValidationStatus.PENDING: (100, 100, 100),
    ValidationStatus.PROCESSING: (30, 100, 200),
    ValidationStatus.ERROR: (200, 40, 40),
}


class ReportGenerator:

    def __init__(self):
        os.makedirs(settings.REPORTS_DIR, exist_ok=True)

    def generate_pdf(self, certificates: List[Certificate], filename: str) -> str:
        """Generate a PDF report."""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph,
                Spacer, Image as RLImage, HRFlowable, PageBreak,
            )
            from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        except ImportError:
            raise RuntimeError("reportlab not installed")

        filepath = os.path.join(settings.REPORTS_DIR, filename)
        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A4),
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=15*mm,
            bottomMargin=15*mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title", parent=styles["Title"],
            fontSize=16, textColor=colors.HexColor("#1a2e4a"),
            spaceAfter=6,
        )
        subtitle_style = ParagraphStyle(
            "Subtitle", parent=styles["Normal"],
            fontSize=9, textColor=colors.HexColor("#666666"),
        )
        cell_style = ParagraphStyle(
            "Cell", parent=styles["Normal"],
            fontSize=7, leading=10,
        )

        story = []

        # Header
        story.append(Paragraph("ICCBPO — Background Certificate Validation Report", title_style))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')} | Total: {len(certificates)} certificates",
            subtitle_style,
        ))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1a2e4a")))
        story.append(Spacer(1, 8*mm))

        # Summary stats
        stats = self._compute_stats(certificates)
        summary_data = [
            ["Metric", "Count", "Percentage"],
            ["Total Certificates", str(stats["total"]), "100%"],
            ["✓ Verified Authentic", str(stats["verified"]),
             f"{stats['verified']/stats['total']*100:.1f}%" if stats["total"] else "0%"],
            ["✗ Failed / Fraudulent", str(stats["failed"]),
             f"{stats['failed']/stats['total']*100:.1f}%" if stats["total"] else "0%"],
            ["⚠ Technical Issue", str(stats["technical"]),
             f"{stats['technical']/stats['total']*100:.1f}%" if stats["total"] else "0%"],
        ]
        summary_table = Table(summary_data, colWidths=[180, 60, 80])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2e4a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f8f9fa"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TEXTCOLOR", (0, 2), (0, 2), colors.HexColor("#155724")),
            ("TEXTCOLOR", (0, 3), (0, 3), colors.HexColor("#721c24")),
            ("TEXTCOLOR", (0, 4), (0, 4), colors.HexColor("#856404")),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 8*mm))

        # Main table
        headers = [
            "File Name", "Holder Name", "Cert Number", "Country",
            "QR URL", "Status", "Confidence", "Issue Date",
            "Screenshot", "Error / Notes", "Uploaded At",
        ]
        col_widths = [90, 80, 65, 55, 100, 85, 50, 50, 55, 90, 70]

        table_data = [headers]
        for cert in certificates:
            status_label = STATUS_LABELS.get(cert.status, cert.status)
            conf_pct = f"{cert.confidence_score*100:.0f}%" if cert.confidence_score else "—"
            qr_url = (cert.qr_url or "")[:60] + ("..." if cert.qr_url and len(cert.qr_url) > 60 else "")
            screenshot = "View" if cert.screenshot_url else "—"

            row = [
                Paragraph(cert.original_filename[:40], cell_style),
                Paragraph(cert.holder_name or "—", cell_style),
                cert.cert_number or "—",
                cert.country or "—",
                Paragraph(qr_url or "No QR", cell_style),
                Paragraph(status_label, cell_style),
                conf_pct,
                cert.issue_date or "—",
                screenshot,
                Paragraph((cert.error_details or cert.analyst_notes or "—")[:80], cell_style),
                cert.uploaded_at.strftime("%Y-%m-%d") if cert.uploaded_at else "—",
            ]
            table_data.append(row)

        main_table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Build style commands with status colors
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a2e4a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#dee2e6")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f8f9fa"), colors.white]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]

        # Color status column
        for i, cert in enumerate(certificates, 1):
            rgb = STATUS_COLORS_RGB.get(cert.status, (100, 100, 100))
            r, g, b = [x/255 for x in rgb]
            c = colors.Color(r, g, b)
            style_cmds.append(("TEXTCOLOR", (5, i), (5, i), c))
            style_cmds.append(("FONTNAME", (5, i), (5, i), "Helvetica-Bold"))

        main_table.setStyle(TableStyle(style_cmds))
        story.append(main_table)

        # Footer
        story.append(Spacer(1, 10*mm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#dee2e6")))
        story.append(Paragraph(
            "CONFIDENTIAL — For authorized ICCBPO HR/Compliance personnel only. "
            "This report was generated automatically. Results should be reviewed by a qualified analyst.",
            ParagraphStyle("Footer", parent=styles["Normal"], fontSize=6,
                          textColor=colors.HexColor("#999999")),
        ))

        doc.build(story)
        logger.info(f"PDF report generated: {filepath}")
        return filepath

    def generate_excel(self, certificates: List[Certificate], filename: str) -> str:
        """Generate an Excel report."""
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, Fill, PatternFill, Alignment, Border, Side, colors as xl_colors
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl not installed")

        filepath = os.path.join(settings.REPORTS_DIR, filename)
        wb = openpyxl.Workbook()

        # ── Summary Sheet ──────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = "ICCBPO Certificate Validation Report"
        ws_summary["A1"].font = Font(bold=True, size=14, color="1A2E4A")
        ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary["A2"].font = Font(size=9, color="666666")
        ws_summary.merge_cells("A1:F1")

        stats = self._compute_stats(certificates)
        summary_rows = [
            ("", ""),
            ("Total Certificates", stats["total"]),
            ("Verified Authentic", stats["verified"]),
            ("Failed / Fraudulent", stats["failed"]),
            ("Technical Issue", stats["technical"]),
            ("Pending", stats["pending"]),
            ("Average Confidence", f"{stats['avg_confidence']*100:.1f}%"),
        ]
        for i, (label, value) in enumerate(summary_rows, 4):
            ws_summary[f"A{i}"] = label
            ws_summary[f"B{i}"] = value
            ws_summary[f"A{i}"].font = Font(bold=True)

        # ── Certificates Sheet ─────────────────────────────────────
        ws = wb.create_sheet("Certificates")

        headers = [
            "File Name", "Holder Name", "Holder ID", "Cert Number",
            "Country", "Language", "Cert Type", "Issue Date", "Expiry Date",
            "Issuing Authority", "QR Found", "QR URL",
            "Status", "Confidence %", "Official Domain",
            "Valid Keywords", "Invalid Keywords", "Fraud Score",
            "Fraud Indicators", "Error Details", "Screenshot URL",
            "Processing Time (s)", "Analyst Notes", "Uploaded At", "Processed At",
        ]

        # Header row
        header_fill = PatternFill("solid", fgColor="1A2E4A")
        header_font = Font(bold=True, color="FFFFFF", size=9)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Status fills
        status_fills = {
            ValidationStatus.VERIFIED_AUTHENTIC: PatternFill("solid", fgColor="C6EFCE"),
            ValidationStatus.FAILED_FRAUDULENT: PatternFill("solid", fgColor="FFC7CE"),
            ValidationStatus.TECHNICAL_ISSUE: PatternFill("solid", fgColor="FFEB9C"),
            ValidationStatus.PENDING: PatternFill("solid", fgColor="EEEEEE"),
            ValidationStatus.PROCESSING: PatternFill("solid", fgColor="BDD7EE"),
            ValidationStatus.ERROR: PatternFill("solid", fgColor="FFC7CE"),
        }

        for row_num, cert in enumerate(certificates, 2):
            row_fill = status_fills.get(cert.status)
            data = [
                cert.original_filename,
                cert.holder_name or "",
                cert.holder_id or "",
                cert.cert_number or "",
                cert.country or "",
                cert.language_detected or "",
                cert.cert_type.value if cert.cert_type else "",
                cert.issue_date or "",
                cert.expiry_date or "",
                cert.issuing_authority or "",
                "Yes" if cert.qr_code_found else "No",
                cert.qr_url or "",
                STATUS_LABELS.get(cert.status, str(cert.status)),
                round(cert.confidence_score * 100, 1),
                "Yes" if cert.is_official_domain else ("No" if cert.is_official_domain is False else ""),
                ", ".join((cert.fraud_indicators or {}).get("valid_keywords", [])) if cert.fraud_indicators else "",
                ", ".join((cert.fraud_indicators or {}).get("invalid_keywords", [])) if cert.fraud_indicators else "",
                round(cert.fraud_score * 100, 1),
                "Possible fraud" if cert.is_potentially_fraudulent else "Clean",
                cert.error_details or "",
                cert.screenshot_url or "",
                round(cert.processing_time_seconds, 2) if cert.processing_time_seconds else "",
                cert.analyst_notes or "",
                cert.uploaded_at.strftime("%Y-%m-%d %H:%M") if cert.uploaded_at else "",
                cert.processed_at.strftime("%Y-%m-%d %H:%M") if cert.processed_at else "",
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(size=8)
                if row_fill:
                    cell.fill = row_fill

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["L"].width = 50
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        wb.save(filepath)
        logger.info(f"Excel report generated: {filepath}")
        return filepath

    def generate_csv(self, certificates: List[Certificate], filename: str) -> str:
        """Generate a CSV report."""
        filepath = os.path.join(settings.REPORTS_DIR, filename)

        fieldnames = [
            "file_name", "holder_name", "holder_id", "cert_number",
            "country", "language", "cert_type", "issue_date",
            "qr_found", "qr_url", "status", "confidence_pct",
            "official_domain", "fraud_score", "is_potentially_fraudulent",
            "error_details", "screenshot_url", "analyst_notes",
            "uploaded_at", "processed_at",
        ]

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for cert in certificates:
                writer.writerow({
                    "file_name": cert.original_filename,
                    "holder_name": cert.holder_name or "",
                    "holder_id": cert.holder_id or "",
                    "cert_number": cert.cert_number or "",
                    "country": cert.country or "",
                    "language": cert.language_detected or "",
                    "cert_type": cert.cert_type.value if cert.cert_type else "",
                    "issue_date": cert.issue_date or "",
                    "qr_found": "Yes" if cert.qr_code_found else "No",
                    "qr_url": cert.qr_url or "",
                    "status": STATUS_LABELS.get(cert.status, str(cert.status)),
                    "confidence_pct": f"{cert.confidence_score*100:.1f}",
                    "official_domain": cert.is_official_domain,
                    "fraud_score": f"{cert.fraud_score*100:.1f}",
                    "is_potentially_fraudulent": cert.is_potentially_fraudulent,
                    "error_details": cert.error_details or "",
                    "screenshot_url": cert.screenshot_url or "",
                    "analyst_notes": cert.analyst_notes or "",
                    "uploaded_at": cert.uploaded_at.isoformat() if cert.uploaded_at else "",
                    "processed_at": cert.processed_at.isoformat() if cert.processed_at else "",
                })

        logger.info(f"CSV report generated: {filepath}")
        return filepath

    def _compute_stats(self, certificates: List[Certificate]) -> dict:
        total = len(certificates)
        verified = sum(1 for c in certificates if c.status == ValidationStatus.VERIFIED_AUTHENTIC)
        failed = sum(1 for c in certificates if c.status == ValidationStatus.FAILED_FRAUDULENT)
        technical = sum(1 for c in certificates if c.status == ValidationStatus.TECHNICAL_ISSUE)
        pending = sum(1 for c in certificates if c.status in (ValidationStatus.PENDING, ValidationStatus.PROCESSING))
        scores = [c.confidence_score for c in certificates if c.confidence_score > 0]
        return {
            "total": total,
            "verified": verified,
            "failed": failed,
            "technical": technical,
            "pending": pending,
            "avg_confidence": sum(scores) / len(scores) if scores else 0.0,
        }
